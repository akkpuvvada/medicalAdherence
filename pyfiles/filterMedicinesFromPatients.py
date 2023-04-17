# Filter out medicines that has patients
import pandas as pd
import openpyxl

patient_file = openpyxl.load_workbook(filename="../results2/filter_patients.xlsx", read_only=True)
patient_file_worksheet = patient_file.active

print(patient_file_worksheet.max_row)
MAX_ROW = patient_file_worksheet.max_row
# MAX_ROW = 20

patients_array = []

for i in range(1, MAX_ROW + 1):
    print(i)
    cell_obj = patient_file_worksheet.cell(row = i, column = 1)
    patients_array.append(cell_obj.value)

print(patients_array)
filter_medicine_list = openpyxl.load_workbook(filename="../results2/filter_medicines2.xlsx", read_only=True)
filter_medicine_list_worksheet = filter_medicine_list.active
print(filter_medicine_list_worksheet.max_row)

# Load the rows
rows = filter_medicine_list_worksheet.rows
headers = [cell.value for cell in next(rows)]
PATIENTS_COL = headers.index('DUPERSID')
# print(patients_array)
data = []
idx = 1
for row in rows:
    print('med',idx)
    record = {}
    patient_exist = False
    # print(row[PATIENTS_COL].value, row[PATIENTS_COL].value in patients_array)
    if row[PATIENTS_COL].value in patients_array:
      for key, cell in zip(headers, row):
        record[key] = cell.value
      data.append(record)
    idx += 1

filter_medicine_list.close()

# Convert to a df
df = pd.DataFrame(data)

# Save to file
df.to_excel("../results2/filter_medicines_3.xlsx")

patient_file.close()