# Filter out list that has only diabetes related medicine
import pandas as pd
import openpyxl

medicine_pre_final = openpyxl.load_workbook(filename="../results2/filter_medicines3.xlsx", read_only=True)
medicine_pre_final_worksheet = medicine_pre_final.active

patient_file = openpyxl.load_workbook(filename="../results2/patients.xlsx", read_only=True)
patient_file_worksheet = patient_file.active
MAX_ROW = patient_file_worksheet.max_row
# MAX_ROW = 5

patients_array = []

# for i in range(1, 10):
for i in range(1, MAX_ROW + 1):
    cell_obj = patient_file_worksheet.cell(row = i, column = 1)
    patients_array.append(cell_obj.value)
# patients_array = ['2320675102']

# print(patients_array)
medicineRows = medicine_pre_final_worksheet.rows
headers = [cell.value for cell in next(medicineRows)]
# print('headers', medicineRows)

patient_idx = headers.index('DUPERSID')
quant_idx = headers.index('RXQUANTY')
med_name = headers.index('RXNAME')
days_up = headers.index('RXDAYSUP')
monthStartIdx = headers.index('RXBEGMM')

medicinesList = ['GLIPIZIDE', 'GLYBURIDE', 'GLICLAZIDE', 'GLIMEPIRIDE', 'TOLBUTAMIDE', 'REPAGLINIDE', 'NATEGLINIDE', 'METFORMIN', 'ROSIGLITAZONE', 'PIOGLITAZONE', 'ACARBOSE', 'MIGLITOL', 'VOGLIBOSE', 'SITAGLIPTIN', 'SAXAGLIPTIN', 'VILDAGLIPTIN', 'LINAGLIPTIN', 'ALOGLIPTIN', 'DAPAGLIFLOZIN', 'CANAGLIFLOZIN', 'EMPAGLIFLOZIN']

data = []
idx = 0

for patient in patients_array:
  print(patient)
  daysUP = [0 for i in range(len(medicinesList))] 
  TotalMedsTobeTaken = [0 for i in range(len(medicinesList))] 
  medPrescribed = [0 for i in range(len(medicinesList))]
  medDosage = [0 for i in range(len(medicinesList))]
  adherence = [0 for i in range(len(medicinesList))]
  record = {}
  id = 0  
  count = 0
  overallAdherence = 0

  # print(medicineRows)
  medicineRows1 = medicine_pre_final_worksheet.rows
  headers = [cell.value for cell in next(medicineRows1)]
  for row in medicineRows1:
    # print('in 1')
    idx += 1
    # print(row[patient_idx].value, patient)
    # print(type(row[patient_idx].value), type(patient))
    if row[patient_idx].value == patient:
      print('inside')
      medicineIndex = medicinesList.index(row[med_name].value)
      # print('medicineIndex', row[days_up].value)
      daysUP[medicineIndex] += row[days_up].value
      # TotalMeds[medicineIndex] += row[quant_idx].value
      if medPrescribed[medicineIndex] != 0 or medPrescribed[medicineIndex] < row[quant_idx].value:
          TotalMedsTobeTaken[medicineIndex] = row[quant_idx].value * (13-row[monthStartIdx].value)
          medDosage[medicineIndex] = row[quant_idx].value / row[days_up].value

  for med_row in medicinesList:
    # print('in 2')
    no_of_medicines = 0
    if daysUP[id] > 0:
      daysCovered = daysUP[id] * medDosage[id]
      print(daysCovered)
      medsTobeTaken = TotalMedsTobeTaken[id]
      print('meds to be taken')
      adherenceCalculated =  daysCovered / medsTobeTaken
      adherence[id] = daysCovered / medsTobeTaken
    id += 1

  for value in adherence:
    # print('in 3')
    if value > 0:
      count += 1
      overallAdherence += value
      print(count, overallAdherence)
  if count > 0:
    # print('in 4')
    overallAdherence = overallAdherence / count
  record['DUPERSID'] = patient
  record['MA.2018'] = overallAdherence
  # print('in herrhr')
  data.append(record)


# Convert to a df
df = pd.DataFrame(data)
print(df)
df.to_excel("../intermediateResults/finalResult1.xlsx")

medicine_pre_final.close()
patient_file.close()
