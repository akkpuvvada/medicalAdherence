# Filter out list that has only diabetes related medicine
import pandas as pd
import openpyxl

medicine_file = openpyxl.load_workbook(filename="../results2/filter_medicines_3_dupers.xlsx", read_only=True)
medicine_file_worksheet = medicine_file.active

print(medicine_file_worksheet.max_row)
MAX_ROW = medicine_file_worksheet.max_row
# MAX_ROW = 20

patients_array = []

for i in range(1, MAX_ROW + 1):
    print(i)
    cell_obj = medicine_file_worksheet.cell(row = i, column = 1)
    patients_array.append(cell_obj.value)

print(patients_array)
patient_file = openpyxl.load_workbook(filename="../results2/filter_output3.xlsx", read_only=True)
patient_file_worksheet = patient_file.active
print(patient_file_worksheet.max_row)

# Load the rows
rows = patient_file_worksheet.rows
headers = [cell.value for cell in next(rows)]
PATIENTS_COL = headers.index('DUPERSID')
# print(patients_array)
data = []
idx = 1
for row in rows:
    print('med',idx)
    record = {}
    patient_exist = False
    # print(row[PATIENTS_COL].value, row[PATIENTS_COL].value in patients_array)
    if row[PATIENTS_COL].value in patients_array:
      for key, cell in zip(headers, row):
        record[key] = cell.value
      data.append(record)
    idx += 1

medicine_file.close()

# Convert to a df
df = pd.DataFrame(data)

# Save to file
df.to_excel("../results2/filter_output4.xlsx")

patient_file.close()